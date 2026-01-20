# Copyright (c) 2019, Frappe Technologies and contributors
# For license information, please see license.txt


import csv
import io
import json
import re
from datetime import date, datetime

import frappe
import mt940
import openpyxl
from frappe import _
from frappe.core.doctype.data_import.data_import import DataImport
from frappe.core.doctype.data_import.importer import Importer, ImportFile
from frappe.utils.background_jobs import enqueue
from frappe.utils.file_manager import get_file, save_file
from frappe.utils.xlsxutils import ILLEGAL_CHARACTERS_RE, handle_html
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

INVALID_VALUES = ("", None)


class BankStatementImport(DataImport):
	# begin: auto-generated types
	# This code is auto-generated. Do not modify anything in this block.

	from typing import TYPE_CHECKING

	if TYPE_CHECKING:
		from frappe.types import DF

		bank: DF.Link | None
		bank_account: DF.Link
		company: DF.Link
		custom_delimiters: DF.Check
		delimiter_options: DF.Data | None
		google_sheets_url: DF.Data | None
		import_file: DF.Attach | None
		import_mt940_fromat: DF.Check
		import_type: DF.Literal["", "Insert New Records", "Update Existing Records"]
		mute_emails: DF.Check
		reference_doctype: DF.Link
		show_failed_logs: DF.Check
		status: DF.Literal["Pending", "Success", "Partial Success", "Error"]
		submit_after_import: DF.Check
		template_options: DF.Code | None
		template_warnings: DF.Code | None
		use_csv_sniffer: DF.Check
	# end: auto-generated types

	def __init__(self, *args, **kwargs):
		super().__init__(*args, **kwargs)

	def validate(self):
		doc_before_save = self.get_doc_before_save()
		if (
			not (self.import_file or self.google_sheets_url)
			or (doc_before_save and doc_before_save.import_file != self.import_file)
			or (doc_before_save and doc_before_save.google_sheets_url != self.google_sheets_url)
		):
			template_options_dict = {}
			column_to_field_map = {}
			bank = frappe.get_doc("Bank", self.bank)
			for i in bank.bank_transaction_mapping:
				column_to_field_map[i.file_field] = i.bank_transaction_field
			template_options_dict["column_to_field_map"] = column_to_field_map
			self.template_options = json.dumps(template_options_dict)

			self.template_warnings = ""

		if self.import_file and not self.import_file.lower().endswith(".txt"):
			self.validate_import_file()
			self.validate_google_sheets_url()

	def start_import(self):
		preview = frappe.get_doc("Bank Statement Import", self.name).get_preview_from_template(
			self.import_file, self.google_sheets_url
		)

		if "Bank Account" not in json.dumps(preview["columns"]):
			frappe.throw(_("Please add the Bank Account column"))

		from frappe.utils.background_jobs import is_job_enqueued
		from frappe.utils.scheduler import is_scheduler_inactive

		run_now = frappe.in_test or frappe.conf.developer_mode
		if is_scheduler_inactive() and not run_now:
			frappe.throw(_("Scheduler is inactive. Cannot import data."), title=_("Scheduler Inactive"))

		job_id = f"bank_statement_import::{self.name}"
		if not is_job_enqueued(job_id):
			enqueue(
				start_import,
				queue="default",
				timeout=6000,
				event="data_import",
				job_id=job_id,
				data_import=self.name,
				bank_account=self.bank_account,
				import_file_path=self.import_file,
				google_sheets_url=self.google_sheets_url,
				bank=self.bank,
				template_options=self.template_options,
				now=run_now,
			)
			return job_id

		return None


def preprocess_mt940_content(content: str) -> str:
	"""Preprocess MT940 content to fix statement number format issues.

	The MT940 standard expects statement numbers to be maximum 5 digits,
	but some banks provide longer statement numbers that cause parsing errors.
	This function truncates statement numbers longer than 5 digits to the last 5 digits.
	"""
	# Fast-path: bail if no :28C: tag exists
	if ":28C:" not in content:
		return content

	# Match :28C: at start of line, capture digits and optional /seq, preserve whitespace
	pattern = re.compile(r"(?m)^(:28C:)(\d{6,})(/\d+)?(\s*)$")

	def replace_statement_number(match):
		prefix = match.group(1)  # ':28C:'
		statement_num = match.group(2)  # The statement number
		sequence_part = match.group(3) or ""  # The sequence part like '/1'
		trailing_space = match.group(4) or ""  # Preserve trailing whitespace

		# If statement number is longer than 5 digits, truncate to last 5 digits
		if len(statement_num) > 5:
			statement_num = statement_num[-5:]

		return prefix + statement_num + sequence_part + trailing_space

	# Apply the replacement
	processed_content = pattern.sub(replace_statement_number, content)
	return processed_content


@frappe.whitelist()
def convert_mt940_to_csv(data_import, mt940_file_path):
	doc = frappe.get_doc("Bank Statement Import", data_import)

	_file_doc, content = get_file(mt940_file_path)

	is_mt940 = is_mt940_format(content)
	if not is_mt940:
		frappe.throw(_("The uploaded file does not appear to be in valid MT940 format."))

	if is_mt940 and not doc.import_mt940_fromat:
		frappe.throw(_("MT940 file detected. Please enable 'Import MT940 Format' to proceed."))

	try:
		# Preprocess MT940 content to fix statement number format issues
		processed_content = preprocess_mt940_content(content)
		transactions = mt940.parse(processed_content)
	except Exception as e:
		frappe.throw(_("Failed to parse MT940 format. Error: {0}").format(str(e)))

	if not transactions:
		frappe.throw(_("Parsed file is not in valid MT940 format or contains no transactions."))

	# Use in-memory file buffer instead of writing to temp file
	csv_buffer = io.StringIO()
	writer = csv.writer(csv_buffer)

	headers = ["Date", "Deposit", "Withdrawal", "Description", "Reference Number", "Bank Account", "Currency"]
	writer.writerow(headers)

	for txn in transactions:
		txn_date = getattr(txn, "date", None)
		raw_date = txn.data.get("date", "")

		if txn_date:
			date_str = txn_date.strftime("%Y-%m-%d")
		elif isinstance(raw_date, date | datetime):
			date_str = raw_date.strftime("%Y-%m-%d")
		else:
			date_str = str(raw_date)

		raw_amount = str(txn.data.get("amount", ""))
		parts = raw_amount.strip().split()
		amount_value = float(parts[0]) if parts else 0.0

		deposit = amount_value if amount_value > 0 else ""
		withdrawal = abs(amount_value) if amount_value < 0 else ""
		description = txn.data.get("extra_details") or ""
		reference = txn.data.get("transaction_reference") or ""
		currency = txn.data.get("currency", "")

		writer.writerow([date_str, deposit, withdrawal, description, reference, doc.bank_account, currency])

	# Prepare in-memory CSV for upload
	csv_content = csv_buffer.getvalue().encode("utf-8")
	csv_buffer.close()

	filename = f"{frappe.utils.now_datetime().strftime('%Y%m%d%H%M%S')}_converted_mt940.csv"

	# Save to File Manager
	saved_file = save_file(filename, csv_content, doc.doctype, doc.name, is_private=True, df="import_file")

	return saved_file.file_url


@frappe.whitelist()
def get_preview_from_template(data_import, import_file=None, google_sheets_url=None):
	return frappe.get_doc("Bank Statement Import", data_import).get_preview_from_template(
		import_file, google_sheets_url
	)


@frappe.whitelist()
def form_start_import(data_import):
	job_id = frappe.get_doc("Bank Statement Import", data_import).start_import()
	return job_id is not None


@frappe.whitelist()
def download_errored_template(data_import_name):
	data_import = frappe.get_doc("Bank Statement Import", data_import_name)
	data_import.export_errored_rows()


@frappe.whitelist()
def download_import_log(data_import_name):
	return frappe.get_doc("Bank Statement Import", data_import_name).download_import_log()


def is_mt940_format(content: str) -> bool:
	"""Check if the content has key MT940 tags"""
	required_tags = [":20:", ":25:", ":28C:", ":61:"]
	return all(tag in content for tag in required_tags)


def parse_data_from_template(raw_data):
	data = []

	for _i, row in enumerate(raw_data):
		if all(v in INVALID_VALUES for v in row):
			# empty row
			continue

		data.append(row)

	return data


def start_import(data_import, bank_account, import_file_path, google_sheets_url, bank, template_options):
	"""This method runs in background job"""

	update_mapping_db(bank, template_options)

	data_import = frappe.get_doc("Bank Statement Import", data_import)
	file = import_file_path if import_file_path else google_sheets_url

	import_file = ImportFile("Bank Transaction", file=file, import_type="Insert New Records")

	data = parse_data_from_template(import_file.raw_data)
	# Importer expects 'Data Import' class, which has 'payload_count' attribute
	if not data_import.get("payload_count"):
		data_import.payload_count = len(data) - 1

	if import_file_path:
		add_bank_account(data, bank_account)
		write_files(import_file, data)

	try:
		i = Importer(data_import.reference_doctype, data_import=data_import)
		i.import_data()
	except Exception:
		frappe.db.rollback()
		data_import.db_set("status", "Error")
		data_import.log_error("Bank Statement Import failed")
	finally:
		frappe.flags.in_import = False

	frappe.publish_realtime("data_import_refresh", {"data_import": data_import.name})


def update_mapping_db(bank, template_options):
	"""Update bank transaction mapping database with template options."""
	bank = frappe.get_doc("Bank", bank)
	for d in bank.bank_transaction_mapping:
		d.delete()

	for d in json.loads(template_options)["column_to_field_map"].items():
		bank.append("bank_transaction_mapping", {"bank_transaction_field": d[1], "file_field": d[0]})

	bank.save()


def add_bank_account(data, bank_account):
	"""Add bank account information to data rows."""
	bank_account_loc = None
	if "Bank Account" not in data[0]:
		data[0].append("Bank Account")
	else:
		for loc, header in enumerate(data[0]):
			if header == "Bank Account":
				bank_account_loc = loc

	for row in data[1:]:
		if bank_account_loc:
			row[bank_account_loc] = bank_account
		else:
			row.append(bank_account)


def write_files(import_file, data):
	"""Write processed data to CSV or Excel files."""
	full_file_path = import_file.file_doc.get_full_path()
	parts = import_file.file_doc.get_extension()
	extension = parts[1]
	extension = extension.lstrip(".")

	if extension == "csv":
		with open(full_file_path, "w", newline="") as file:
			writer = csv.writer(file)
			writer.writerows(data)
	elif extension in ("xlsx", "xls"):
		write_xlsx(data, "trans", file_path=full_file_path)


def write_xlsx(data, sheet_name, wb=None, column_widths=None, file_path=None):
	"""Write data to Excel file with formatting."""
	# from xlsx utils with changes
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook(write_only=True)

	ws = wb.create_sheet(sheet_name, 0)

	for i, column_width in enumerate(column_widths):
		if column_width:
			ws.column_dimensions[get_column_letter(i + 1)].width = column_width

	row1 = ws.row_dimensions[1]
	row1.font = Font(name="Calibri", bold=True)

	for row in data:
		clean_row = []
		for item in row:
			if isinstance(item, str) and (sheet_name not in ["Data Import Template", "Data Export"]):
				value = handle_html(item)
			else:
				value = item

			if isinstance(item, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
				# Remove illegal characters from the string
				value = re.sub(ILLEGAL_CHARACTERS_RE, "", value)

			clean_row.append(value)

		ws.append(clean_row)

	wb.save(file_path)
	return True


@frappe.whitelist()
def get_import_status(docname):
	import_status = {}

	data_import = frappe.get_doc("Bank Statement Import", docname)
	import_status["status"] = data_import.status

	logs = frappe.get_all(
		"Data Import Log",
		fields=[{"COUNT": "*", "as": "count"}, "success"],
		filters={"data_import": docname},
		group_by="success",
	)

	total_payload_count = 0

	for log in logs:
		total_payload_count += log.get("count", 0)
		if log.get("success"):
			import_status["success"] = log.get("count")
		else:
			import_status["failed"] = log.get("count")

	import_status["total_records"] = total_payload_count

	return import_status


@frappe.whitelist()
def get_import_logs(docname: str):
	frappe.has_permission("Bank Statement Import", throw=True)

	return frappe.get_all(
		"Data Import Log",
		fields=["success", "docname", "messages", "exception", "row_indexes"],
		filters={"data_import": docname},
		limit_page_length=5000,
		order_by="log_index",
	)


@frappe.whitelist()
def upload_bank_statement(**args):
	args = frappe._dict(args)
	bsi = frappe.new_doc("Bank Statement Import")

	if args.company:
		bsi.update(
			{
				"company": args.company,
			}
		)

	if args.bank_account:
		bsi.update({"bank_account": args.bank_account})

	return bsi
